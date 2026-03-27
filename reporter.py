import os
import csv
import shutil
import json
import logging
import html as html_mod

import matcher

logger = logging.getLogger(__name__)

AUDIT_DIR      = "audit_package"
MATCHED_DIR    = os.path.join(AUDIT_DIR, "matched")
SUMMARY_XLSX   = os.path.join(AUDIT_DIR, "summary.xlsx")
SUMMARY_HTML   = os.path.join(AUDIT_DIR, "summary.html")
OVERRIDES_FILE = "overrides.json"


def configure(cfg):
    global AUDIT_DIR, MATCHED_DIR, SUMMARY_XLSX, SUMMARY_HTML, OVERRIDES_FILE
    audit = cfg.get("audit_dir", AUDIT_DIR)
    AUDIT_DIR      = audit
    MATCHED_DIR    = os.path.join(audit, "matched")
    SUMMARY_XLSX   = os.path.join(audit, "summary.xlsx")
    SUMMARY_HTML   = os.path.join(audit, "summary.html")
    OVERRIDES_FILE = cfg.get("overrides_file", OVERRIDES_FILE)


def _file_url(path):
    """Return a file:// URL for the given path (Windows-compatible)."""
    return "file:///" + os.path.abspath(path).replace("\\", "/")


def _fmt_fields(vscore, date_ok, batch=False):
    parts = ["amount=BATCH" if batch else "amount=OK"]
    if vscore >= matcher._STRONG_VENDOR:
        parts.append(f"vendor=OK(score={vscore})")
    elif vscore >= matcher._FUZZY_VENDOR:
        parts.append(f"vendor=FUZZY(score={vscore})")
    else:
        parts.append(f"vendor=MISS(score={vscore})")
    parts.append("date=OK" if date_ok else "date=MISS")
    return "  ".join(parts)


def _print_preview(matched, unmatched_txns, orphan_invoices):
    total_txns = len(matched) + len(unmatched_txns)
    sep = "=" * 70

    logger.info("\n%s", sep)
    logger.info("MATCHED QB TRANSACTIONS (%d of %d)", len(matched), total_txns)
    logger.info(sep)
    for m in matched:
        txn = m["transaction"]
        tag = ("VENDOR-A-RECEIPT " if m.get("vendor_a_receipt")
               else "VENDOR-B-RECEIPT " if m.get("vendor_b_receipt")
               else "BATCH " if m["batch"] else "")
        logger.info("  [%s] %sTXN %s | %s | $%.2f | %s",
                    m["confidence"], tag, txn["id"], txn["date"],
                    txn["amount"], txn["vendor"] or "(no vendor)")
        for inv in m["invoices"]:
            fields = _fmt_fields(m.get("vscore", 0), m.get("date_ok", False), m["batch"])
            logger.info("    -> %s", os.path.basename(inv["file"]))
            logger.info("       inv: $%.2f | %s | %s",
                        inv["amount"], inv["date"] or "no date", inv["vendor"] or "no vendor")
            logger.info("       matched: %s", fields)

    logger.info("\n%s", sep)
    logger.info("UNMATCHED QB TRANSACTIONS (%d of %d)", len(unmatched_txns), total_txns)
    logger.info(sep)
    for r in unmatched_txns:
        txn = r["transaction"]
        logger.info("  TXN %s | %s | $%.2f | %s",
                    txn["id"], txn["date"], txn["amount"], txn["vendor"] or "(no vendor)")
        logger.info("    why: %s", r["diag"])

    logger.info("\n%s", sep)
    logger.info("ORPHAN INVOICES (%d PDFs with no QB match)", len(orphan_invoices))
    logger.info(sep)
    for inv in orphan_invoices[:20]:
        amt = f"${inv['amount']:.2f}" if inv["amount"] is not None else "$None"
        logger.info("  %s: %s | %s | %s",
                    os.path.basename(inv["file"]), amt,
                    inv["date"] or "no date", inv["vendor"] or "no vendor")
    if len(orphan_invoices) > 20:
        logger.info("  ... (%d more not shown)", len(orphan_invoices) - 20)

    logger.info("\n%s", sep)
    logger.info("AUDIT PACKAGE SUMMARY")
    logger.info(sep)
    logger.info("  QB transactions total    : %d", total_txns)
    logger.info("  Matched (have invoice)   : %d", len(matched))
    logger.info("  Unmatched (review queue) : %d", len(unmatched_txns))
    logger.info("  Orphan invoices (skipped): %d", len(orphan_invoices))
    logger.info("  audit_package/matched/   — %d invoice file(s)",
                sum(len(m["invoices"]) for m in matched))


def _load_overrides():
    """Load overrides.json keyed by transaction ID. Returns {} if file absent."""
    if not os.path.exists(OVERRIDES_FILE):
        return {}
    with open(OVERRIDES_FILE, encoding="utf-8") as f:
        return json.load(f)


def _unmatched_note(txn):
    """Return a default note for an unmatched transaction."""
    return "Invoice not found — manual review required"


def _write_html_summary(matched, unmatched_txns, overrides=None):
    if overrides is None:
        overrides = {}

    total        = len(matched) + len(unmatched_txns)
    manual_count = sum(
        1 for r in unmatched_txns
        if overrides.get(r["transaction"]["id"], {}).get("attachments")
    )
    matched_count   = len(matched) + manual_count
    unmatched_count = len(unmatched_txns) - manual_count
    rate = matched_count / total * 100 if total else 0

    def esc(s):
        return html_mod.escape(str(s)) if s is not None else ""

    def _links_html(file_list):
        return "".join(
            f'<a href="matched/{html_mod.escape(os.path.basename(f))}">'
            f'{html_mod.escape(os.path.basename(f))}</a>'
            for f in file_list
        )

    rows_html = []

    for m in matched:
        txn = m["transaction"]
        ov  = overrides.get(txn["id"], {})
        vendor    = txn.get("vendor") or (m["invoices"][0].get("vendor") if m["invoices"] else "")
        inv_files = [] if ov.get("replace") else [inv["file"] for inv in m["invoices"]]
        seen = {os.path.basename(f) for f in inv_files}
        for a in ov.get("attachments", []):
            if os.path.basename(a) not in seen:
                inv_files.append(a)
                seen.add(os.path.basename(a))
        links = _links_html(inv_files)
        note  = esc(ov.get("note", ""))
        rows_html.append(
            f'<tr class="matched" data-id="{esc(txn["id"])}">'
            f'<td>matched</td>'
            f'<td>{esc(txn["date"])}</td>'
            f'<td>{esc(vendor)}</td>'
            f'<td class="amount">${txn["amount"]:,.2f}</td>'
            f'<td class="txnid">{esc(txn["id"])}</td>'
            f'<td class="links">{links}</td>'
            f'<td>{note}</td>'
            f'</tr>'
        )

    for r in unmatched_txns:
        txn         = r["transaction"]
        ov          = overrides.get(txn["id"], {})
        attachments = ov.get("attachments", [])
        note        = esc(ov.get("note") or _unmatched_note(txn))
        if attachments:
            status = "manual"
            links  = _links_html(attachments)
        else:
            status = "unmatched"
            links  = ""
        rows_html.append(
            f'<tr class="{status}" data-id="{esc(txn["id"])}">'
            f'<td>{status}</td>'
            f'<td>{esc(txn["date"])}</td>'
            f'<td>{esc(txn.get("vendor") or "")}</td>'
            f'<td class="amount">${txn["amount"]:,.2f}</td>'
            f'<td class="txnid">{esc(txn["id"])}</td>'
            f'<td class="links">{links}</td>'
            f'<td>{note}</td>'
            f'</tr>'
        )

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>Invoice Audit Summary</title>
<style>
  body {{ font-family: Arial, sans-serif; font-size: 13px; margin: 20px; }}
  h2 {{ margin-bottom: 6px; }}
  .summary {{ background: #f0f0f0; padding: 10px 14px; border-radius: 4px;
              margin-bottom: 14px; font-size: 14px; }}
  table {{ border-collapse: collapse; width: 100%; }}
  th, td {{ border: 1px solid #ccc; padding: 6px 10px; text-align: left; vertical-align: top; }}
  th {{ background: #1f4e79; color: #fff; cursor: pointer; user-select: none; white-space: nowrap; }}
  th.sort-asc:after  {{ content: " \u25b2"; font-size: 10px; }}
  th.sort-desc:after {{ content: " \u25bc"; font-size: 10px; }}
  th:not(.sort-asc):not(.sort-desc):after {{ content: " \u25b4\u25be"; font-size: 10px; color: #aad4f5; }}
  tr.matched   td {{ background: #e2efda; }}
  tr.unmatched td {{ background: #fce4d6; }}
  tr.manual    td {{ background: #fff2cc; }}
  tr:hover td {{ filter: brightness(0.95); }}
  td.amount {{ text-align: right; font-variant-numeric: tabular-nums; }}
  td.links a {{ display: block; color: #0563c1; word-break: break-all; }}
</style>
</head>
<body>
<h2>Invoice Audit Summary</h2>
<div class="summary">
  Total QB transactions: <strong>{total}</strong> &nbsp;|&nbsp;
  Matched: <strong>{matched_count}</strong> &nbsp;|&nbsp;
  Unmatched: <strong>{unmatched_count}</strong> &nbsp;|&nbsp;
  Manual: <strong>{manual_count}</strong> &nbsp;|&nbsp;
  Match rate: <strong>{rate:.1f}%</strong>
</div>
<table id="auditTable">
<thead>
<tr>
  <th onclick="sortTable(0)">Status</th>
  <th onclick="sortTable(1)">Date</th>
  <th onclick="sortTable(2)">Vendor</th>
  <th onclick="sortTable(3)">Amount</th>
  <th onclick="sortTable(4)">Transaction ID</th>
  <th onclick="sortTable(5)">Invoice Filename</th>
  <th onclick="sortTable(6)">Notes</th>
</tr>
</thead>
<tbody>
{"".join(rows_html)}
</tbody>
</table>
<script>
var _sortCol = -1, _sortAsc = true;

function cellVal(cell, col) {{
  var t = cell.innerText.trim();
  if (col === 4) {{
    var m = t.match(/(\\d+)$/);
    return m ? parseInt(m[1], 10) : 0;
  }}
  if (col === 3) {{
    var n = parseFloat(t.replace(/[$,]/g, ""));
    return isNaN(n) ? 0 : n;
  }}
  return t;
}}

function sortTable(col) {{
  var tbl = document.getElementById("auditTable");
  var tbody = tbl.tBodies[0];
  var rows = Array.from(tbody.rows);
  var asc = (_sortCol === col) ? !_sortAsc : true;
  rows.sort(function(a, b) {{
    var x = cellVal(a.cells[col], col);
    var y = cellVal(b.cells[col], col);
    if (typeof x === "number") return asc ? x - y : y - x;
    return asc ? x.localeCompare(y) : y.localeCompare(x);
  }});
  rows.forEach(function(r) {{ tbody.appendChild(r); }});
  Array.from(tbl.tHead.rows[0].cells).forEach(function(th, i) {{
    th.classList.remove("sort-asc", "sort-desc");
    if (i === col) th.classList.add(asc ? "sort-asc" : "sort-desc");
  }});
  _sortCol = col; _sortAsc = asc;
}}

// Default: sort by date descending on load
sortTable(1); sortTable(1);
</script>
</body>
</html>"""

    os.makedirs(AUDIT_DIR, exist_ok=True)
    with open(SUMMARY_HTML, "w", encoding="utf-8") as f:
        f.write(html)


def _write_audit_package(matched, unmatched_txns, orphan_invoices):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    os.makedirs(MATCHED_DIR, exist_ok=True)

    overrides = _load_overrides()

    # Copy all override attachments into matched/ so HTML links resolve
    for txn_id, ov in overrides.items():
        for src in ov.get("attachments", []):
            if os.path.isfile(src):
                shutil.copy2(src, os.path.join(MATCHED_DIR, os.path.basename(src)))
            else:
                logger.warning("[override] attachment not found, skipping: %s", src)

    rows = []

    for m in matched:
        txn = m["transaction"]
        ov  = overrides.get(txn["id"], {})
        inv_files = []
        if not ov.get("replace"):
            for inv in m["invoices"]:
                dest = os.path.join(MATCHED_DIR, os.path.basename(inv["file"]))
                shutil.copy2(inv["file"], dest)
                inv_files.append((os.path.basename(inv["file"]), _file_url(dest)))
        seen_fnames = {f for f, _ in inv_files}
        for src in ov.get("attachments", []):
            fname = os.path.basename(src)
            if fname not in seen_fnames:
                dest = os.path.join(MATCHED_DIR, fname)
                inv_files.append((fname, _file_url(dest)))
                seen_fnames.add(fname)
        first_fname, first_url = inv_files[0] if inv_files else ("", "")
        all_fnames = "; ".join(f for f, _ in inv_files)
        vendor = txn.get("vendor") or (m["invoices"][0].get("vendor") if m["invoices"] else "")
        rows.append({
            "status":           "matched",
            "date":             txn["date"],
            "vendor":           vendor,
            "amount":           txn["amount"],
            "transaction_id":   txn["id"],
            "invoice_filename": all_fnames,
            "local_file_path":  first_fname,
            "_link_url":        first_url,
            "notes":            ov.get("note", ""),
        })

    for r in unmatched_txns:
        txn         = r["transaction"]
        ov          = overrides.get(txn["id"], {})
        attachments = ov.get("attachments", [])
        if attachments:
            status      = "manual"
            first_fname = os.path.basename(attachments[0])
            first_url   = _file_url(os.path.join(MATCHED_DIR, first_fname))
            all_fnames  = "; ".join(os.path.basename(a) for a in attachments)
        else:
            status = "unmatched"
            first_fname = first_url = all_fnames = ""
        rows.append({
            "status":           status,
            "date":             txn["date"],
            "vendor":           txn.get("vendor") or "",
            "amount":           txn["amount"],
            "transaction_id":   txn["id"],
            "invoice_filename": all_fnames,
            "local_file_path":  first_fname,
            "_link_url":        first_url,
            "notes":            ov.get("note") or _unmatched_note(txn),
        })

    # --- Write XLSX ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice Summary"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    display_headers = ["Status", "Date", "Vendor", "Amount", "Transaction ID",
                       "Invoice Filename", "Invoice Link", "Notes"]
    for col, h in enumerate(display_headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    matched_fill   = PatternFill("solid", fgColor="E2EFDA")
    unmatched_fill = PatternFill("solid", fgColor="FCE4D6")
    manual_fill    = PatternFill("solid", fgColor="FFF2CC")
    link_font      = Font(color="0563C1", underline="single")
    link_col       = display_headers.index("Invoice Link") + 1
    _fill_map      = {"matched": matched_fill, "manual": manual_fill, "unmatched": unmatched_fill}

    for row_idx, row in enumerate(rows, 2):
        fill   = _fill_map.get(row["status"], unmatched_fill)
        values = [
            row["status"], row["date"], row["vendor"], row["amount"],
            row["transaction_id"], row["invoice_filename"],
            None,
            row["notes"],
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = fill
            if col == 4:
                cell.number_format = '$#,##0.00'
        link_cell = ws.cell(row=row_idx, column=link_col)
        link_cell.fill = fill
        if row.get("_link_url"):
            url     = row["_link_url"].replace('"', '""')
            display = row["local_file_path"].replace('"', '""')
            link_cell.value = f'=HYPERLINK("{url}","{display}")'
            link_cell.font  = link_font

    col_widths = [10, 12, 24, 12, 16, 45, 20, 45]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    wb.save(SUMMARY_XLSX)

    # --- HTML summary ---
    _write_html_summary(matched, unmatched_txns, overrides=overrides)

    # --- Orphan invoice log ---
    orphan_log = os.path.join(AUDIT_DIR, "orphan_invoices.csv")
    with open(orphan_log, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["filename", "amount", "date", "vendor"])
        writer.writeheader()
        for inv in orphan_invoices:
            writer.writerow({
                "filename": os.path.basename(inv["file"]),
                "amount":   inv["amount"],
                "date":     inv["date"] or "",
                "vendor":   inv["vendor"] or "",
            })

    total_txns = len(matched) + len(unmatched_txns)
    logger.info("Audit package written to %s/", AUDIT_DIR)
    logger.info("  QB transactions total    : %d", total_txns)
    logger.info("  Matched (have invoice)   : %d", len(matched))
    logger.info("  Unmatched (review queue) : %d", len(unmatched_txns))
    logger.info("  matched/  : %d invoice file(s)", sum(len(m["invoices"]) for m in matched))
    logger.info("  summary.xlsx: %d row(s)", len(rows))
    logger.info("  orphan_invoices.csv: %d invoice(s) with no QB match", len(orphan_invoices))
