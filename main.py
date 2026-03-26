import os
import re
import json
import csv
import shutil
import argparse
from datetime import date, datetime
from itertools import combinations

from qb import get_transactions
from extract import extract_invoice_data
from qb_attach import upload_file, attach_to_transaction

try:
    from thefuzz import fuzz
    _FUZZY = True
except ImportError:
    _FUZZY = False

INVOICE_DIR = "invoices"
AUDIT_DIR   = "audit_package"
MATCHED_DIR = os.path.join(AUDIT_DIR, "matched")
SUMMARY_XLSX = os.path.join(AUDIT_DIR, "summary.xlsx")

# Vendor fuzzy-match thresholds
_STRONG_VENDOR = 80   # contributes to High confidence
_FUZZY_VENDOR  = 50   # contributes to Medium confidence

# ---------------------------------------------------------------------------
# Vendor configuration
# ---------------------------------------------------------------------------
# QB transactions whose payee name matches one of these require an invoice
# with a passing vendor score before they can be matched.
# Add your own vendor names here.
_KNOWN_QB_VENDOR_RE = re.compile(
    r"\b(Vendor A|Vendor B|Vendor C|Coffee Supplier|Bakery Co)\b",
    re.IGNORECASE,
)

# Vendor aliases: if a vendor appears under multiple names, normalize them
# to a single canonical string before fuzzy matching.
# Format: (pattern, canonical_name)
_VENDOR_ALIASES = [
    # Example: "Vendor B LLC" and "VendorB" both normalize to "Vendor B"
    (re.compile(r"vendor\s*b", re.IGNORECASE), "Vendor B"),
]


def _txn_has_known_vendor(txn):
    target = f"{txn.get('vendor') or ''} {txn.get('memo') or ''}"
    return bool(_KNOWN_QB_VENDOR_RE.search(target))


_CONF_RANK = {"High": 3, "Medium": 2, "Low": 1}


def _apply_vendor_aliases(inv_vendor, filename):
    """Return inv_vendor rewritten to its canonical name if an alias matches."""
    text = f"{inv_vendor or ''} {os.path.basename(filename)}"
    for pattern, canonical in _VENDOR_ALIASES:
        if pattern.search(text):
            return canonical
    return inv_vendor


# ---------------------------------------------------------------------------
# Matching helpers
# ---------------------------------------------------------------------------

def _amount_ok(inv_amount, txn_amount, tolerance=False):
    """Return (matches, exact).
    tolerance=True: also accept txn up to 3.5% above invoice (e.g. processing fees).
    """
    if inv_amount is None:
        return False, False
    if abs(inv_amount - txn_amount) <= 0.01:
        return True, True
    if inv_amount > 0 and abs(txn_amount - inv_amount) / inv_amount <= 0.005:
        return True, False
    if tolerance and inv_amount > 0 and 0 < (txn_amount - inv_amount) / inv_amount <= 0.035:
        return True, False
    return False, False


def _date_ok(inv_date, txn_date, max_days=60):
    """Invoice date must be 0–max_days before the QB transaction date."""
    if not inv_date or not txn_date:
        return False
    try:
        inv = date.fromisoformat(inv_date)
        txn = date.fromisoformat(txn_date)
        delta = (txn - inv).days
        return 0 <= delta <= max_days
    except (ValueError, TypeError):
        return False


def _vendor_score(inv_vendor, txn_vendor, txn_memo):
    """Return 0–100 fuzzy match score against QB vendor name."""
    if not inv_vendor or not _FUZZY:
        return 0
    targets = [txn_vendor] if txn_vendor else ([txn_memo] if txn_memo else [])
    if not targets:
        return 0
    return max(fuzz.partial_ratio(inv_vendor.lower(), t.lower()) for t in targets)


def _confidence(vscore, date_matches):
    """
    High   = strong vendor match AND date in range
    Medium = date in range OR fuzzy vendor match
    Low    = neither (amount-only match)
    """
    if vscore >= _STRONG_VENDOR and date_matches:
        return "High"
    if date_matches or vscore >= _FUZZY_VENDOR:
        return "Medium"
    return "Low"


# ---------------------------------------------------------------------------
# Diagnostics
# ---------------------------------------------------------------------------

def _diagnose_unmatched(inv, transactions):
    """Return a short string explaining why inv could not be matched."""
    if inv["amount"] is None:
        return "no amount extracted"
    exact = [t for t in transactions if _amount_ok(inv["amount"], t["amount"])[0]]
    if exact:
        parts = []
        for t in exact[:3]:
            do = _date_ok(inv["date"], t["date"])
            vs = _vendor_score(inv["vendor"], t["vendor"], t["memo"])
            issues = []
            if not do:
                if not inv["date"]:
                    issues.append("no date")
                else:
                    try:
                        delta = (date.fromisoformat(t["date"]) - date.fromisoformat(inv["date"])).days
                        issues.append(f"date delta={delta}d")
                    except (ValueError, TypeError):
                        issues.append("unparseable date")
            if vs < _FUZZY_VENDOR:
                issues.append(f"vendor score={vs}<{_FUZZY_VENDOR}")
            label = ", ".join(issues) if issues else "already claimed"
            parts.append(f"TXN {t['id']} ({t['vendor'] or 'no vendor'}): {label}")
        return "exact amount hit — " + "; ".join(parts)
    near = sorted(
        [t for t in transactions if abs(t["amount"] - inv["amount"]) / max(t["amount"], 0.01) <= 0.30],
        key=lambda t: abs(t["amount"] - inv["amount"]),
    )
    if near:
        t = near[0]
        diff_pct = (inv["amount"] - t["amount"]) / t["amount"] * 100
        return f"no exact amount match; nearest TXN {t['id']} ${t['amount']:.2f} ({diff_pct:+.1f}%)"
    return "no transaction within 30% of this amount"


def _diagnose_unmatched_txn(txn, invoices):
    """Return a short string explaining why a QB transaction could not be matched."""
    amt = txn["amount"]
    exact = [inv for inv in invoices if inv["amount"] is not None and abs(inv["amount"] - amt) <= 0.01]
    if exact:
        parts = []
        for inv in exact[:3]:
            eff = _apply_vendor_aliases(inv.get("vendor"), inv.get("file", ""))
            vs = _vendor_score(eff, txn["vendor"], txn["memo"])
            do = _date_ok(inv["date"], txn["date"])
            issues = []
            if vs < _FUZZY_VENDOR:
                issues.append(f"vendor mismatch (inv={eff or '?'}, score={vs})")
            if not do:
                issues.append("date out of range")
            parts.append(", ".join(issues) if issues else "already claimed by another txn")
        return "exact invoice amount exists — " + "; ".join(parts)
    near = sorted(
        [inv for inv in invoices
         if inv["amount"] is not None
         and abs(inv["amount"] - amt) / max(amt, 0.01) <= 0.30],
        key=lambda inv: abs(inv["amount"] - amt),
    )
    if near:
        inv = near[0]
        diff_pct = (inv["amount"] - amt) / amt * 100
        eff = _apply_vendor_aliases(inv.get("vendor"), inv.get("file", ""))
        return f"nearest invoice ${inv['amount']:.2f} ({diff_pct:+.1f}%) — {eff or os.path.basename(inv['file'])}"
    return "no invoice within 30% of this amount"


# ---------------------------------------------------------------------------
# Vendor-specific receipt pre-passes
# ---------------------------------------------------------------------------
# Some vendors send a single consolidated payment that covers multiple invoices.
# For these vendors, a payments JSON file maps each payment to its invoice numbers
# so the matcher can correctly link multiple PDFs to one QB transaction.
#
# Format of vendor_a_payments.json:
# [
#   {
#     "date": "2025-03-05",
#     "total": 887.45,
#     "invoices": [
#       {"invoice_number": "INV-001", "amount": 249.99},
#       {"invoice_number": "INV-002", "amount": 637.46}
#     ]
#   }, ...
# ]
#
# If you don't have a vendor that sends consolidated payments, these pre-passes
# are skipped automatically when the JSON file is absent.
# ---------------------------------------------------------------------------

VENDOR_A_PAYMENTS_FILE = "vendor_a_payments.json"
VENDOR_B_PAYMENTS_FILE = "vendor_b_payments.json"

# Pattern to extract an invoice number from a PDF filename.
# Customize this regex to match your vendor's filename convention.
_VENDOR_A_INV_NUM_RE = re.compile(r"INV-(\d+)", re.IGNORECASE)


def _load_vendor_a_receipts():
    """Load vendor_a_payments.json. Returns (inv_to_receipt dict, receipts list)."""
    if not os.path.exists(VENDOR_A_PAYMENTS_FILE):
        return {}, []
    with open(VENDOR_A_PAYMENTS_FILE, encoding="utf-8") as f:
        receipts = json.load(f)
    inv_to_receipt = {}
    for r in receipts:
        for entry in r.get("invoices", []):
            num = entry.get("invoice_number") or entry
            inv_to_receipt[str(num)] = r
    return inv_to_receipt, receipts


def _is_vendor_a_invoice(inv):
    """True if this PDF is from Vendor A (by vendor name or filename pattern)."""
    vendor = (inv.get("vendor") or "").lower()
    fname  = os.path.basename(inv.get("file", "")).lower()
    return "vendor a" in vendor or fname.startswith("vendor_a_")


def _load_vendor_b_receipts():
    if not os.path.exists(VENDOR_B_PAYMENTS_FILE):
        return []
    with open(VENDOR_B_PAYMENTS_FILE, encoding="utf-8") as f:
        return json.load(f)


def _is_vendor_b_invoice(inv):
    """True if this PDF is from Vendor B."""
    eff = _apply_vendor_aliases(inv.get("vendor"), inv.get("file", ""))
    return (eff or "").lower() == "vendor b"


# ---------------------------------------------------------------------------
# Core matching
# ---------------------------------------------------------------------------

def _match_invoices(transactions, invoices):
    """
    Returns:
        matched        — list of {transaction, invoices, confidence, batch}
        unmatched_txns — list of {transaction, diag}
        orphan_invoices — list of invoice dicts with no QB match
    """
    used = set()   # indices into `invoices` that have been claimed
    matched = []

    # ---- Vendor A receipt pre-pass ----
    # Strategy: match receipt total → QB transaction (±$0.01),
    # then attach the individual invoice PDFs for that payment.
    _inv_to_receipt, receipts = _load_vendor_a_receipts()
    if receipts:
        used_txn_ids = set()
        for r in receipts:
            if r["total"] is None:
                continue
            best_txn = next(
                (txn for txn in transactions
                 if txn["id"] not in used_txn_ids
                 and abs(txn["amount"] - r["total"]) <= 0.01),
                None,
            )
            if best_txn is None:
                continue

            avail = [(i, inv) for i, inv in enumerate(invoices)
                     if i not in used and _is_vendor_a_invoice(inv)]

            # Try to match individual invoice PDFs by amount
            receipt_amts = [e["amount"] for e in r.get("invoices", []) if e.get("amount") is not None]
            pool = list(avail)
            matched_by_amt = []
            for r_amt in receipt_amts:
                hit = next(
                    ((i, inv) for i, inv in pool
                     if inv["amount"] is not None and abs(inv["amount"] - r_amt) <= 0.01),
                    None,
                )
                if hit:
                    matched_by_amt.append(hit)
                    pool = [(i, inv) for i, inv in pool if i != hit[0]]

            if len(matched_by_amt) == len(receipt_amts) and matched_by_amt:
                group = matched_by_amt
                conf  = "High"
            else:
                # Fall back to date proximity (within 14 days of receipt date)
                conf  = "Medium"
                group = []
                if r.get("date"):
                    try:
                        r_date = date.fromisoformat(r["date"])
                        group  = [
                            (i, inv) for i, inv in avail
                            if inv.get("date")
                            and abs((date.fromisoformat(inv["date"]) - r_date).days) <= 14
                        ]
                    except (ValueError, TypeError):
                        pass

            if not group:
                continue

            matched.append({
                "transaction": best_txn,
                "invoices":    [inv for _, inv in group],
                "confidence":  conf,
                "batch":       len(group) > 1,
                "vscore":      100,
                "date_ok":     _date_ok(r["date"], best_txn["date"]),
                "vendor_a_receipt": True,
            })
            used.update(i for i, _ in group)
            used_txn_ids.add(best_txn["id"])

    # ---- Vendor B receipt pre-pass ----
    # Strategy: match receipt amount → QB Vendor B transaction (±$0.01, ±5 days),
    # then attach all Vendor B invoice PDFs dated within 14 days of the receipt.
    vendor_b_receipts = sorted(_load_vendor_b_receipts(), key=lambda r: r.get("date") or "")
    if vendor_b_receipts:
        receipt_dates = []
        for r in vendor_b_receipts:
            try:
                receipt_dates.append(date.fromisoformat(r["date"]) if r.get("date") else None)
            except (ValueError, TypeError):
                receipt_dates.append(None)

        # Pre-assign each Vendor B invoice to its closest receipt by date
        inv_best_receipt = {}
        for i, inv in enumerate(invoices):
            if i in used or not _is_vendor_b_invoice(inv) or not inv.get("date"):
                continue
            try:
                inv_date = date.fromisoformat(inv["date"])
            except (ValueError, TypeError):
                continue
            best_ri, best_delta = None, float("inf")
            for ri, rd in enumerate(receipt_dates):
                if rd is None:
                    continue
                delta = abs((inv_date - rd).days)
                if delta <= 14 and delta < best_delta:
                    best_delta = delta
                    best_ri    = ri
            if best_ri is not None:
                inv_best_receipt[i] = best_ri

        used_txn_ids_b = set()
        for ri, r in enumerate(vendor_b_receipts):
            if not r.get("amount"):
                continue
            r_date_parsed = None
            try:
                r_date_parsed = date.fromisoformat(r["date"]) if r.get("date") else None
            except (ValueError, TypeError):
                pass

            best_txn = None
            for txn in transactions:
                if txn["id"] in used_txn_ids_b:
                    continue
                if abs(txn["amount"] - r["amount"]) <= 0.01:
                    if r_date_parsed and txn.get("date"):
                        try:
                            if abs((date.fromisoformat(txn["date"]) - r_date_parsed).days) > 5:
                                continue
                        except (ValueError, TypeError):
                            pass
                    best_txn = txn
                    break
            if best_txn is None:
                continue

            group = [(i, inv) for i, inv in enumerate(invoices)
                     if i not in used and inv_best_receipt.get(i) == ri]
            if not group:
                continue

            # Prefer exact-amount match; fall back to nearest within tolerance
            near_group = [(i, inv) for i, inv in group
                          if inv.get("amount") is not None
                          and _amount_ok(inv["amount"], r["amount"], tolerance=True)[0]]
            if near_group:
                exact = [(i, inv) for i, inv in near_group
                         if abs(inv["amount"] - r["amount"]) <= 0.01]
                group = (exact or near_group)[:1]
            else:
                if len(group) == 1:
                    inv_amt = group[0][1].get("amount") or 0
                    if inv_amt > 0 and abs(inv_amt - r["amount"]) / r["amount"] > 0.10:
                        continue

            matched.append({
                "transaction": best_txn,
                "invoices":    [inv for _, inv in group],
                "confidence":  "Medium",
                "batch":       len(group) > 1,
                "vscore":      100,
                "date_ok":     _date_ok(r["date"], best_txn["date"]),
                "vendor_b_receipt": True,
            })
            used.update(i for i, _ in group)
            used_txn_ids_b.add(best_txn["id"])

    # ---- General single-invoice pass ----
    # Process oldest→newest so each transaction claims its chronologically
    # closest invoice before a later transaction can take it.
    already_matched_ids = {m["transaction"]["id"] for m in matched}
    for txn in sorted(transactions, key=lambda t: t["date"] or ""):
        if txn["id"] in already_matched_ids:
            continue

        best_idx = best_conf = None
        best_vs = 0
        best_do = False
        best_is_invoice = False

        for i, inv in enumerate(invoices):
            if i in used:
                continue
            use_tolerance = _is_vendor_b_invoice(inv)
            amt_ok, amt_exact = _amount_ok(inv["amount"], txn["amount"], tolerance=use_tolerance)
            if not amt_ok:
                continue
            eff_vendor = _apply_vendor_aliases(inv["vendor"], inv["file"])
            vs   = _vendor_score(eff_vendor, txn["vendor"], txn["memo"])
            if _txn_has_known_vendor(txn) and vs < _FUZZY_VENDOR:
                continue
            do   = _date_ok(inv["date"], txn["date"], max_days=60)
            conf = _confidence(vs, do)
            if not amt_exact:
                conf = min(conf, "Medium", key=lambda c: _CONF_RANK[c])
            is_invoice = inv.get("document_type", "invoice") != "sales_order"
            better = (
                best_conf is None
                or _CONF_RANK[conf] > _CONF_RANK[best_conf]
                or (_CONF_RANK[conf] == _CONF_RANK[best_conf] and is_invoice and not best_is_invoice)
            )
            if better:
                best_idx, best_conf, best_vs, best_do, best_is_invoice = i, conf, vs, do, is_invoice

        if best_idx is not None:
            matched.append({
                "transaction": txn,
                "invoices":    [invoices[best_idx]],
                "confidence":  best_conf,
                "batch":       False,
                "vscore":      best_vs,
                "date_ok":     best_do,
            })
            used.add(best_idx)
            continue

        # ---- Batch fallback (2–3 invoices that sum to transaction total) ----
        available = [
            (i, inv) for i, inv in enumerate(invoices)
            if i not in used
            and inv["amount"] is not None
            and 0 < inv["amount"] < txn["amount"]
        ]
        found_batch = False
        for size in (2, 3):
            for combo in combinations(available, size):
                idxs = [idx for idx, _ in combo]
                invs = [inv for _, inv in combo]
                # All invoices must share the same vendor
                vendors = {(inv["vendor"] or "").strip().lower() for inv in invs}
                if len(vendors) > 1 or vendors == {""}:
                    continue
                # All invoice dates within 7 days of each other
                try:
                    dates = [date.fromisoformat(inv["date"]) for inv in invs if inv["date"]]
                except (ValueError, TypeError):
                    continue
                if len(dates) != size or (max(dates) - min(dates)).days > 7:
                    continue
                if _txn_has_known_vendor(txn):
                    batch_vs = _vendor_score(
                        _apply_vendor_aliases(invs[0]["vendor"], invs[0]["file"]),
                        txn["vendor"], txn["memo"],
                    )
                    if batch_vs < _FUZZY_VENDOR:
                        continue
                if abs(sum(inv["amount"] for inv in invs) - txn["amount"]) <= 0.01:
                    matched.append({
                        "transaction": txn,
                        "invoices":    invs,
                        "confidence":  "Medium",
                        "batch":       True,
                    })
                    used.update(idxs)
                    found_batch = True
                    break
            if found_batch:
                break

    # ---- Unmatched QB transactions ----
    matched_txn_ids = {m["transaction"]["id"] for m in matched}
    unmatched_txns  = [
        {"transaction": t, "diag": _diagnose_unmatched_txn(t, invoices)}
        for t in transactions
        if t["id"] not in matched_txn_ids
    ]

    # ---- Orphan invoices (PDFs with no QB match) ----
    orphan_invoices = [inv for i, inv in enumerate(invoices) if i not in used]

    return matched, unmatched_txns, orphan_invoices


# ---------------------------------------------------------------------------
# Audit package helpers
# ---------------------------------------------------------------------------

def _file_url(path):
    """Return a file:// URL for the given path (Windows-compatible)."""
    return "file:///" + os.path.abspath(path).replace("\\", "/")


def _fmt_fields(vscore, date_ok, batch=False):
    parts = ["amount=BATCH" if batch else "amount=OK"]
    if vscore >= _STRONG_VENDOR:
        parts.append(f"vendor=OK(score={vscore})")
    elif vscore >= _FUZZY_VENDOR:
        parts.append(f"vendor=FUZZY(score={vscore})")
    else:
        parts.append(f"vendor=MISS(score={vscore})")
    parts.append("date=OK" if date_ok else "date=MISS")
    return "  ".join(parts)


def _print_preview(matched, unmatched_txns, orphan_invoices):
    total_txns = len(matched) + len(unmatched_txns)

    print(f"\n{'='*70}")
    print(f"MATCHED QB TRANSACTIONS ({len(matched)} of {total_txns})")
    print(f"{'='*70}")
    for m in matched:
        txn = m["transaction"]
        tag = ("VENDOR-A-RECEIPT " if m.get("vendor_a_receipt")
               else "VENDOR-B-RECEIPT " if m.get("vendor_b_receipt")
               else "BATCH " if m["batch"] else "")
        print(f"  [{m['confidence']}] {tag}TXN {txn['id']} | {txn['date']} | "
              f"${txn['amount']:.2f} | {txn['vendor'] or '(no vendor)'}")
        for inv in m["invoices"]:
            fields = _fmt_fields(m.get("vscore", 0), m.get("date_ok", False), m["batch"])
            print(f"    -> {os.path.basename(inv['file'])}")
            print(f"       inv: ${inv['amount']:.2f} | {inv['date'] or 'no date'} | "
                  f"{inv['vendor'] or 'no vendor'}")
            print(f"       matched: {fields}")

    print(f"\n{'='*70}")
    print(f"UNMATCHED QB TRANSACTIONS ({len(unmatched_txns)} of {total_txns})")
    print(f"{'='*70}")
    for r in unmatched_txns:
        txn = r["transaction"]
        print(f"  TXN {txn['id']} | {txn['date']} | ${txn['amount']:.2f} | "
              f"{txn['vendor'] or '(no vendor)'}")
        print(f"    why: {r['diag']}")

    print(f"\n{'='*70}")
    print(f"ORPHAN INVOICES ({len(orphan_invoices)} PDFs with no QB match)")
    print(f"{'='*70}")
    for inv in orphan_invoices[:20]:
        amt = f"${inv['amount']:.2f}" if inv["amount"] is not None else "$None"
        print(f"  {os.path.basename(inv['file'])}: {amt} | "
              f"{inv['date'] or 'no date'} | {inv['vendor'] or 'no vendor'}")
    if len(orphan_invoices) > 20:
        print(f"  ... ({len(orphan_invoices) - 20} more not shown)")

    print(f"\n{'='*70}")
    print("AUDIT PACKAGE SUMMARY")
    print(f"{'='*70}")
    print(f"  QB transactions total    : {total_txns}")
    print(f"  Matched (have invoice)   : {len(matched)}")
    print(f"  Unmatched (review queue) : {len(unmatched_txns)}")
    print(f"  Orphan invoices (skipped): {len(orphan_invoices)}")
    print(f"  audit_package/matched/   — {sum(len(m['invoices']) for m in matched)} invoice file(s)")


SUMMARY_HTML    = os.path.join(AUDIT_DIR, "summary.html")
OVERRIDES_FILE  = "overrides.json"


def _load_overrides():
    """Load overrides.json keyed by transaction ID. Returns {} if file absent."""
    if not os.path.exists(OVERRIDES_FILE):
        return {}
    with open(OVERRIDES_FILE, encoding="utf-8") as f:
        return json.load(f)


def _unmatched_note(txn):
    """Return a default note for an unmatched transaction."""
    return "Invoice not found — manual review required"


def _write_html_summary(matched, unmatched_txns, overrides=None):
    import html as html_mod

    if overrides is None:
        overrides = {}

    total        = len(matched) + len(unmatched_txns)
    manual_count = sum(
        1 for r in unmatched_txns
        if overrides.get(r["transaction"]["id"], {}).get("attachments")
    )
    matched_count   = len(matched) + manual_count
    unmatched_count = len(unmatched_txns) - manual_count
    rate = matched_count / total * 100 if total else 0

    def esc(s):
        return html_mod.escape(str(s)) if s is not None else ""

    def _links_html(file_list):
        return "".join(
            f'<a href="matched/{html_mod.escape(os.path.basename(f))}">'
            f'{html_mod.escape(os.path.basename(f))}</a>'
            for f in file_list
        )

    rows_html = []

    for m in matched:
        txn = m["transaction"]
        ov  = overrides.get(txn["id"], {})
        vendor    = txn.get("vendor") or (m["invoices"][0].get("vendor") if m["invoices"] else "")
        inv_files = [] if ov.get("replace") else [inv["file"] for inv in m["invoices"]]
        seen = {os.path.basename(f) for f in inv_files}
        for a in ov.get("attachments", []):
            if os.path.basename(a) not in seen:
                inv_files.append(a)
                seen.add(os.path.basename(a))
        links = _links_html(inv_files)
        note  = esc(ov.get("note", ""))
        rows_html.append(
            f'<tr class="matched" data-id="{esc(txn["id"])}">'
            f'<td>matched</td>'
            f'<td>{esc(txn["date"])}</td>'
            f'<td>{esc(vendor)}</td>'
            f'<td class="amount">${txn["amount"]:,.2f}</td>'
            f'<td class="txnid">{esc(txn["id"])}</td>'
            f'<td class="links">{links}</td>'
            f'<td>{note}</td>'
            f'</tr>'
        )

    for r in unmatched_txns:
        txn         = r["transaction"]
        ov          = overrides.get(txn["id"], {})
        attachments = ov.get("attachments", [])
        note        = esc(ov.get("note") or _unmatched_note(txn))
        if attachments:
            status = "manual"
            links  = _links_html(attachments)
        else:
            status = "unmatched"
            links  = ""
        rows_html.append(
            f'<tr class="{status}" data-id="{esc(txn["id"])}">'
            f'<td>{status}</td>'
            f'<td>{esc(txn["date"])}</td>'
            f'<td>{esc(txn.get("vendor") or "")}</td>'
            f'<td class="amount">${txn["amount"]:,.2f}</td>'
            f'<td class="txnid">{esc(txn["id"])}</td>'
            f'<td class="links">{links}</td>'
            f'<td>{note}</td>'
            f'</tr>'
        )

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>Invoice Audit Summary</title>
<style>
  body {{ font-family: Arial, sans-serif; font-size: 13px; margin: 20px; }}
  h2 {{ margin-bottom: 6px; }}
  .summary {{ background: #f0f0f0; padding: 10px 14px; border-radius: 4px;
              margin-bottom: 14px; font-size: 14px; }}
  table {{ border-collapse: collapse; width: 100%; }}
  th, td {{ border: 1px solid #ccc; padding: 6px 10px; text-align: left; vertical-align: top; }}
  th {{ background: #1f4e79; color: #fff; cursor: pointer; user-select: none; white-space: nowrap; }}
  th.sort-asc:after  {{ content: " \u25b2"; font-size: 10px; }}
  th.sort-desc:after {{ content: " \u25bc"; font-size: 10px; }}
  th:not(.sort-asc):not(.sort-desc):after {{ content: " \u25b4\u25be"; font-size: 10px; color: #aad4f5; }}
  tr.matched   td {{ background: #e2efda; }}
  tr.unmatched td {{ background: #fce4d6; }}
  tr.manual    td {{ background: #fff2cc; }}
  tr:hover td {{ filter: brightness(0.95); }}
  td.amount {{ text-align: right; font-variant-numeric: tabular-nums; }}
  td.links a {{ display: block; color: #0563c1; word-break: break-all; }}
</style>
</head>
<body>
<h2>Invoice Audit Summary</h2>
<div class="summary">
  Total QB transactions: <strong>{total}</strong> &nbsp;|&nbsp;
  Matched: <strong>{matched_count}</strong> &nbsp;|&nbsp;
  Unmatched: <strong>{unmatched_count}</strong> &nbsp;|&nbsp;
  Manual: <strong>{manual_count}</strong> &nbsp;|&nbsp;
  Match rate: <strong>{rate:.1f}%</strong>
</div>
<table id="auditTable">
<thead>
<tr>
  <th onclick="sortTable(0)">Status</th>
  <th onclick="sortTable(1)">Date</th>
  <th onclick="sortTable(2)">Vendor</th>
  <th onclick="sortTable(3)">Amount</th>
  <th onclick="sortTable(4)">Transaction ID</th>
  <th onclick="sortTable(5)">Invoice Filename</th>
  <th onclick="sortTable(6)">Notes</th>
</tr>
</thead>
<tbody>
{"".join(rows_html)}
</tbody>
</table>
<script>
var _sortCol = -1, _sortAsc = true;

function cellVal(cell, col) {{
  var t = cell.innerText.trim();
  if (col === 4) {{
    var m = t.match(/(\\d+)$/);
    return m ? parseInt(m[1], 10) : 0;
  }}
  if (col === 3) {{
    var n = parseFloat(t.replace(/[$,]/g, ""));
    return isNaN(n) ? 0 : n;
  }}
  return t;
}}

function sortTable(col) {{
  var tbl = document.getElementById("auditTable");
  var tbody = tbl.tBodies[0];
  var rows = Array.from(tbody.rows);
  var asc = (_sortCol === col) ? !_sortAsc : true;
  rows.sort(function(a, b) {{
    var x = cellVal(a.cells[col], col);
    var y = cellVal(b.cells[col], col);
    if (typeof x === "number") return asc ? x - y : y - x;
    return asc ? x.localeCompare(y) : y.localeCompare(x);
  }});
  rows.forEach(function(r) {{ tbody.appendChild(r); }});
  Array.from(tbl.tHead.rows[0].cells).forEach(function(th, i) {{
    th.classList.remove("sort-asc", "sort-desc");
    if (i === col) th.classList.add(asc ? "sort-asc" : "sort-desc");
  }});
  _sortCol = col; _sortAsc = asc;
}}

// Default: sort by date descending on load
sortTable(1); sortTable(1);
</script>
</body>
</html>"""

    os.makedirs(AUDIT_DIR, exist_ok=True)
    with open(SUMMARY_HTML, "w", encoding="utf-8") as f:
        f.write(html)


def _write_audit_package(matched, unmatched_txns, orphan_invoices):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    os.makedirs(MATCHED_DIR, exist_ok=True)

    overrides = _load_overrides()

    # Copy all override attachments into matched/ so HTML links resolve
    for txn_id, ov in overrides.items():
        for src in ov.get("attachments", []):
            if os.path.isfile(src):
                shutil.copy2(src, os.path.join(MATCHED_DIR, os.path.basename(src)))
            else:
                print(f"  [override] attachment not found, skipping: {src}")

    rows = []

    for m in matched:
        txn = m["transaction"]
        ov  = overrides.get(txn["id"], {})
        inv_files = []
        if not ov.get("replace"):
            for inv in m["invoices"]:
                dest = os.path.join(MATCHED_DIR, os.path.basename(inv["file"]))
                shutil.copy2(inv["file"], dest)
                inv_files.append((os.path.basename(inv["file"]), _file_url(dest)))
        seen_fnames = {f for f, _ in inv_files}
        for src in ov.get("attachments", []):
            fname = os.path.basename(src)
            if fname not in seen_fnames:
                dest = os.path.join(MATCHED_DIR, fname)
                inv_files.append((fname, _file_url(dest)))
                seen_fnames.add(fname)
        first_fname, first_url = inv_files[0] if inv_files else ("", "")
        all_fnames = "; ".join(f for f, _ in inv_files)
        vendor = txn.get("vendor") or (m["invoices"][0].get("vendor") if m["invoices"] else "")
        rows.append({
            "status":           "matched",
            "date":             txn["date"],
            "vendor":           vendor,
            "amount":           txn["amount"],
            "transaction_id":   txn["id"],
            "invoice_filename": all_fnames,
            "local_file_path":  first_fname,
            "_link_url":        first_url,
            "notes":            ov.get("note", ""),
        })

    for r in unmatched_txns:
        txn         = r["transaction"]
        ov          = overrides.get(txn["id"], {})
        attachments = ov.get("attachments", [])
        if attachments:
            status      = "manual"
            first_fname = os.path.basename(attachments[0])
            first_url   = _file_url(os.path.join(MATCHED_DIR, first_fname))
            all_fnames  = "; ".join(os.path.basename(a) for a in attachments)
        else:
            status = "unmatched"
            first_fname = first_url = all_fnames = ""
        rows.append({
            "status":           status,
            "date":             txn["date"],
            "vendor":           txn.get("vendor") or "",
            "amount":           txn["amount"],
            "transaction_id":   txn["id"],
            "invoice_filename": all_fnames,
            "local_file_path":  first_fname,
            "_link_url":        first_url,
            "notes":            ov.get("note") or _unmatched_note(txn),
        })

    # --- Write XLSX ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice Summary"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    display_headers = ["Status", "Date", "Vendor", "Amount", "Transaction ID",
                       "Invoice Filename", "Invoice Link", "Notes"]
    for col, h in enumerate(display_headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    matched_fill   = PatternFill("solid", fgColor="E2EFDA")
    unmatched_fill = PatternFill("solid", fgColor="FCE4D6")
    manual_fill    = PatternFill("solid", fgColor="FFF2CC")
    link_font      = Font(color="0563C1", underline="single")
    link_col       = display_headers.index("Invoice Link") + 1
    _fill_map      = {"matched": matched_fill, "manual": manual_fill, "unmatched": unmatched_fill}

    for row_idx, row in enumerate(rows, 2):
        fill   = _fill_map.get(row["status"], unmatched_fill)
        values = [
            row["status"], row["date"], row["vendor"], row["amount"],
            row["transaction_id"], row["invoice_filename"],
            None,
            row["notes"],
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = fill
            if col == 4:
                cell.number_format = '$#,##0.00'
        link_cell = ws.cell(row=row_idx, column=link_col)
        link_cell.fill = fill
        if row.get("_link_url"):
            url     = row["_link_url"].replace('"', '""')
            display = row["local_file_path"].replace('"', '""')
            link_cell.value = f'=HYPERLINK("{url}","{display}")'
            link_cell.font  = link_font

    col_widths = [10, 12, 24, 12, 16, 45, 20, 45]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    wb.save(SUMMARY_XLSX)

    # --- HTML summary ---
    _write_html_summary(matched, unmatched_txns, overrides=overrides)

    # --- Orphan invoice log ---
    orphan_log = os.path.join(AUDIT_DIR, "orphan_invoices.csv")
    with open(orphan_log, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["filename", "amount", "date", "vendor"])
        writer.writeheader()
        for inv in orphan_invoices:
            writer.writerow({
                "filename": os.path.basename(inv["file"]),
                "amount":   inv["amount"],
                "date":     inv["date"] or "",
                "vendor":   inv["vendor"] or "",
            })

    total_txns = len(matched) + len(unmatched_txns)
    print(f"\nAudit package written to {AUDIT_DIR}/")
    print(f"  QB transactions total    : {total_txns}")
    print(f"  Matched (have invoice)   : {len(matched)}")
    print(f"  Unmatched (review queue) : {len(unmatched_txns)}")
    print(f"  matched/  : {sum(len(m['invoices']) for m in matched)} invoice file(s)")
    print(f"  summary.xlsx: {len(rows)} row(s)")
    print(f"  orphan_invoices.csv: {len(orphan_invoices)} invoice(s) with no QB match")


# ---------------------------------------------------------------------------
# CSV transaction loader
# ---------------------------------------------------------------------------

QB_CSV_PATH = "transactions.csv"

# Only include transactions in these account categories.
# Update to match your QuickBooks account names.
_INCLUDE_CATEGORIES = {"Cost of Goods Sold", "Supplies"}


def _parse_date(raw):
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m-%d-%Y", "%B %d, %Y", "%b %d, %Y"):
        try:
            return datetime.strptime(raw.strip(), fmt).date().isoformat()
        except ValueError:
            continue
    return ""


def _parse_amount(raw):
    """Strip $, commas, parens; return float or None."""
    raw      = raw.strip().replace(",", "").replace("$", "")
    negative = raw.startswith("(") and raw.endswith(")")
    raw      = raw.strip("()")
    try:
        return -float(raw) if negative else float(raw)
    except ValueError:
        return None


# Column name mappings for both QBO export formats
_COL_MAPS = [
    ("Type",             "Category",         "Total",  "Payee"),   # new format
    ("Transaction type", "Account full name", "Amount", "Name"),    # old format
]


def load_transactions_from_csv(path=QB_CSV_PATH):
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader     = csv.DictReader(f)
        rows       = list(reader)
        fieldnames = reader.fieldnames or []

    col_type = col_cat = col_amount = col_vendor = None
    for type_col, cat_col, amt_col, vendor_col in _COL_MAPS:
        if type_col in fieldnames:
            col_type, col_cat, col_amount, col_vendor = type_col, cat_col, amt_col, vendor_col
            break

    if col_type is None:
        raise ValueError(f"Unrecognised CSV format — headers: {fieldnames}")

    transactions = []
    seq = 1
    for row in rows:
        if row.get(col_type, "").strip() != "Expense":
            continue
        category = row.get(col_cat, "").strip()
        if not any(cat in category for cat in _INCLUDE_CATEGORIES):
            continue
        raw_amount = _parse_amount(row.get(col_amount, ""))
        if raw_amount is None:
            continue
        transactions.append({
            "id":     f"csv_{seq}",
            "date":   _parse_date(row.get("Date", "").strip()),
            "amount": abs(raw_amount),
            "vendor": row.get(col_vendor, "").strip(),
            "memo":   category,
        })
        seq += 1

    print(f"  {len(transactions)} transactions loaded from {path}\n")
    return transactions


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main(dry_run=True, from_csv=False):
    mode_tag = "[DRY RUN]" if dry_run else "[LIVE]"
    src_tag  = "[CSV]"     if from_csv else "[API]"
    print(f"--- Invoice Attach Bot {mode_tag} {src_tag} ---\n")

    print("Loading QuickBooks transactions...")
    if from_csv:
        transactions = load_transactions_from_csv()
    else:
        transactions = get_transactions()
        print(f"  {len(transactions)} transactions loaded\n")

    if not os.path.isdir(INVOICE_DIR):
        print(f"ERROR: {INVOICE_DIR}/ not found. Add your invoice PDFs to that folder.")
        return

    pdf_paths = [
        os.path.join(root, f)
        for root, _, files in os.walk(INVOICE_DIR)
        for f in files
        if f.lower().endswith(".pdf")
    ]
    print(f"Scanning {INVOICE_DIR}/ — {len(pdf_paths)} PDF(s) found\n")

    print("Extracting invoice data...")
    invoices = []
    for path in pdf_paths:
        inv = extract_invoice_data(path)
        invoices.append(inv)
        print(f"  {os.path.basename(path):40s}  "
              f"amount={inv['amount']}  date={inv['date']}  vendor={inv['vendor']}")

    print("\nMatching invoices to transactions...")
    matched, unmatched_txns, orphan_invoices = _match_invoices(transactions, invoices)

    _print_preview(matched, unmatched_txns, orphan_invoices)

    if dry_run:
        print("\n[DRY RUN] Nothing written to disk. No QB calls made.")
        return

    _write_audit_package(matched, unmatched_txns, orphan_invoices)

    print("\nAttaching matched invoices to QuickBooks...")
    for m in matched:
        txn = m["transaction"]
        for inv in m["invoices"]:
            try:
                attachable_id, sync_token = upload_file(inv["file"])
                attach_to_transaction(attachable_id, sync_token, txn["id"])
                print(f"  OK  {os.path.basename(inv['file'])} -> TXN {txn['id']}")
            except Exception as e:
                print(f"  ERR {os.path.basename(inv['file'])}: {e}")

    print("\nDone.")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Invoice Attach Bot")
    parser.add_argument(
        "--from-csv",
        action="store_true",
        help="Load QB transactions from transactions.csv instead of the API",
    )
    parser.add_argument(
        "--live",
        action="store_true",
        help="Write audit package and attach invoices to QB (default: dry run)",
    )
    args = parser.parse_args()
    main(dry_run=not args.live, from_csv=args.from_csv)
